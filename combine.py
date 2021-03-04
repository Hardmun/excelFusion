import os
import sys
from copy import copy

from openpyxl import load_workbook, Workbook

class xls(Workbook):
    pass

"""global path"""
projectDir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(
    os.path.abspath(__file__))

def combineFiles():
    wb1 = load_workbook(os.path.join(projectDir, "test_files", "Client.xlsx"))
    ws1 = wb1.create_sheet("lol")

    wb2 = load_workbook(os.path.join(projectDir, "test_files", "Ярославль_remastered.xlsx"))
    ws2 = wb2.active
    for row in ws2.iter_rows():
        for cell in row:
            try:
                new_cell = ws1.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)
            except:
                asdf = 0
            if cell.has_style:
                pass
                # new_cell.font = cell.font
                # new_cell.border = cell.border
                # new_cell.fill = cell.fill
                # new_cell.number_format = cell.number_format
                # new_cell.protection = cell.protection
                # new_cell.alignment = cell.alignment

                # new_cell._style = copy(cell._style)
        # ws1.append((cell.value for cell in row))

    wb1.save(os.path.join(projectDir, "test_files", "fusion.xlsx"))

def combineFilesByopenpyxl():
    wb1 = Workbook()
    # wb1 = load_workbook(os.path.join(projectDir, "test_files", "Client.xlsx"))

    wb2 = load_workbook(os.path.join(projectDir, "test_files", "Ярославль_remastered.xlsx"))
    ws2 = wb2.active

    # ws1 = wb1.copy_worksheet(ws2)

    wb1.save(os.path.join(projectDir, "test_files", "fusion.xlsx"))

def copySheet(target, source):
    for (row, col), source_cell in source._cells.items():
        target_cell = target.cell(column=col, row=row)
        target_cell._value = source_cell._value
        target_cell.data_type = source_cell.data_type

        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

    for attr in ('row_dimensions', 'column_dimensions'):
        src = getattr(source, attr)
        trg = getattr(target, attr)
        for key, dim in src.items():
            trg[key] = copy(dim)
            trg[key].worksheet = trg

    target.sheet_format = copy(source.sheet_format)
    target.sheet_properties = copy(source.sheet_properties)
    target.merged_cells = copy(source.merged_cells)
    target.page_margins = copy(source.page_margins)
    target.page_setup = copy(source.page_setup)
    target.print_options = copy(source.print_options)

def copyCell():
    wb1 = load_workbook(os.path.join(projectDir, "test_files", "Client.xlsx"))
    target = wb1.create_sheet("lol")

    wb2 = load_workbook(os.path.join(projectDir, "test_files", "Ярославль_remastered.xlsx"))
    source = wb2.active

    copySheet(target=target, source=source)

    wb1.save(os.path.join(projectDir, "test_files", "fusion.xlsx"))

def copySheetfromFile():
    wb1 = load_workbook(os.path.join(projectDir, "test_files", "Ярославль_remastered.xlsx"))
    sheetToCopy = wb1["TDSheet"]
    wb1.copy_worksheet(sheetToCopy)

    # wb2 = load_workbook(os.path.join(projectDir, "test_files", "Ярославль_remastered.xlsx"))
    # ws2 = wb2.active
    wb1.save(os.path.join(projectDir, "test_files", "fusion.xlsx"))

if __name__ == '__main__':
    # copySheetfromFile()
    copyCell()
    # combineFilesByopenpyxl()
    # combineFiles()
