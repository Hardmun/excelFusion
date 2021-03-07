import logging
import os
import sys
from copy import copy

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

projectDir: str = ""
loggerglobal = None

def getGlobalVariables():
    """global path"""
    global projectDir
    projectDir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(
        os.path.abspath(__file__))

    global loggerglobal
    loggerglobal = logging.getLogger("errors")
    loggerglobal.setLevel(logging.ERROR)

    formatter = logging.Formatter("%(asctime)s:%(message)s")

    globalHandler = logging.FileHandler(os.path.join(projectDir, "errors.log"))
    globalHandler.setLevel(logging.raiseExceptions)
    globalHandler.setFormatter(formatter)

    loggerglobal.addHandler(globalHandler)

def logDecorator(func):
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except BaseException as errMsg:
            loggerglobal.exception(f"An error has been occurred in function {func.__name__}", exc_info=errMsg)
            sys.exit(0)

    return wrapper

@logDecorator
def IsItNumber(value_check):
    try:
        conv_value = int(value_check)
        return isinstance(conv_value, int)
    except:
        return False

@logDecorator
def r1c1_to_a1(row: int, column: int, formula: str):
    len_formula = len(formula)
    a1 = ""
    parsing_started = False
    start_Row = False
    currentRow = ""
    start_Col = False
    currentCol = ""

    for i, t in zip(formula, range(0, len(formula))):
        "rows parsing"
        if i.upper() == "R":
            next_symb = formula[t + 1]
            if next_symb.upper() == "C":
                currentRow = row
                parsing_started = True
            elif next_symb == "[" or IsItNumber(next_symb):
                start_Row = True
                parsing_started = True
        elif i == "[" and start_Row:
            pass
        elif i == "]" and start_Row:
            currentRow = row + int(currentRow)
            start_Row = False
        elif i.upper() == "C" and start_Row:
            currentRow = int(currentRow)
            start_Row = False
        elif start_Row:
            currentRow += i

        "column parsing"
        if i.upper() == "C":
            if (t + 1) < len_formula:
                next_symb = formula[t + 1]
                if next_symb == "[" or IsItNumber(next_symb):
                    start_Col = True
                elif not next_symb.isalpha():
                    currentCol = column
            else:
                currentCol = column
        elif i == "[" and start_Col:
            pass
        elif i == "]" and start_Col:
            currentCol = column + int(currentCol)
            start_Col = False
        elif start_Col:
            if (t + 1) < len_formula:
                next_symb = formula[t + 1]
                if not IsItNumber(next_symb):
                    currentCol += i
                    if next_symb != "]":
                        currentCol = int(currentCol)
                        start_Col = False
                else:
                    currentCol += i
            else:
                currentCol += i
                currentCol = int(currentCol)
                start_Col = False

        if not parsing_started:
            a1 += i
        elif isinstance(currentRow, int) and isinstance(currentCol, int):
            a1 += f"{get_column_letter(currentCol)}" \
                  f"{str(currentRow)}"
            currentRow = ""
            currentCol = ""
            parsing_started = False

    return a1

@logDecorator
def copySheet(target, source):
    for (row, col), source_cell in source._cells.items():
        target_cell = target.cell(column=col, row=row)
        target_cell._value = source_cell._value
        target_cell.data_type = source_cell.data_type

        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

        if source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)

        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)

    for attr in ('row_dimensions', 'column_dimensions'):
        src = getattr(source, attr)
        trg = getattr(target, attr)
        for key, dim in src.items():
            trg[key] = copy(dim)
            trg[key].worksheet = trg

    target.sheet_format = copy(source.sheet_format)
    target.sheet_properties = copy(source.sheet_properties)
    target.merged_cells = copy(source.merged_cells)
    target.page_margins = copy(source.page_margins)
    target.page_setup = copy(source.page_setup)
    target.print_options = copy(source.print_options)

@logDecorator
def insertFormulas(sheet):
    for (row, col), cell in sheet._cells.items():
        if cell.comment:
            comment: str = cell.comment.text
            find_separator = comment.find("|")
            deleteComment = True

            if find_separator != -1:
                "cell format"
                cell_Format = comment[:find_separator].replace("format_cell:", "")
                cell.number_format = cell_Format
                "formula"
                formula = comment[-(len(comment) - find_separator - 1):].replace("formula_R1C1:", "")
                cell.value = r1c1_to_a1(row=row, column=col, formula=formula).replace(";", ",")
            elif comment.find("formula_R1C1:") != -1:
                "formula"
                formula = comment.replace("formula_R1C1:", "")
                cell.value = r1c1_to_a1(row=row, column=col, formula=formula).replace(";", ",")
            elif comment.find("format_cell:") != -1:
                "format"
                cell_Format = comment.replace("format_cell:", "")
                cell.number_format = cell_Format

                "converting string to float"
                cellValue = cell.value
                if cellValue is not None:
                    cell.value = float(cellValue.replace(",", ".").replace(" ", ""))

            elif comment.find("formula:") != -1:
                "formula"
                formula = comment.replace("formula:", "")
                cell.value = r1c1_to_a1(row=row, column=col, formula=formula).replace(";", ",")
            else:
                deleteComment = False

            if deleteComment:
                cell.comment = None

@logDecorator
def ExcelFusion(sheet_name, tempDir, fileExcel):
    wb_path = os.path.join(projectDir, tempDir, f'{sheet_name}.xlsx')

    """file exists"""
    if not os.path.isfile(wb_path):
        return f"File: {wb_path} doesn't exist!!!"

    if fileExcel is None:
        fileExcel = load_workbook(wb_path)
        ws = fileExcel.worksheets[0]
        ws.title = sheet_name
        insertFormulas(ws)
    else:
        wb_from = load_workbook(wb_path)
        ws_from = wb_from.worksheets[0]
        insertFormulas(ws_from)
        ws = fileExcel.create_sheet(title=sheet_name, index=0)
        copySheet(ws, ws_from)

    return fileExcel

@logDecorator
def readFiles(fileSettings_string):
    """convert to dict"""
    fileSettings = eval(fileSettings_string)

    """checking key - files"""
    settings = fileSettings.get("settings")

    tempDir = settings.get("uuid")

    files = settings.get("files")
    if files is not None:
        wb = None
        for curr_file in files:
            if curr_file:
                wb = ExcelFusion(curr_file, tempDir, wb)

        saveAs = os.path.join(projectDir, settings.get("uuid"), "fusion.xlsx")
        wb.save(saveAs)

if __name__ == '__main__':
    getGlobalVariables()
    if len(sys.argv) == 2:
        files_to_read = sys.argv[1]
        readFiles(files_to_read)
    else:
        loggerglobal.exception(f"Wrong parameters: {str(sys.argv)}")
